"""
Problem 1: Deterministic Optimization via Lagrangian Relaxation + Plot-level DP
Two scenarios: (1) excess wasted, (2) excess sold at 50% discount
"""

import numpy as np
from copy import deepcopy
from config import *
from data_loader import preprocess_all, get_yield_cost_price

# ─── Global data (set after preprocessing) ───
_plots = None
_compat = None
_yield_data = None
_baseline_sales = None
_legume_init = None


def set_global_data(data):
    global _plots, _compat, _yield_data, _baseline_sales, _legume_init, _ycp_cache
    _plots = data['plots']
    _compat = data['compat']
    _yield_data = data['yield_data']
    _baseline_sales = data['baseline']['production']
    _legume_init = data['legume_status']

    # ─── Precompute yield/cost/price cache ───
    # This avoids ~1.3亿 function calls in the DP inner loops.
    # Key: (crop_id, land_type, season) -> (yield, cost, price)
    # Land types seen in data: 平旱地, 梯田, 山坡地, 水浇地, 普通大棚, 智慧大棚
    # Seasons: 单季, 第一季, 第二季
    _ycp_cache = {}
    land_types = ['平旱地', '梯田', '山坡地', '水浇地', '普通大棚', '智慧大棚']
    seasons = ['单季', SEASON1, SEASON2]
    for cid in ALL_CROPS:
        for lt in land_types:
            for season in seasons:
                key = (cid, lt, season)
                yld, cost, price = get_yield_cost_price(_yield_data, cid, lt, season)
                if yld > 0:
                    _ycp_cache[key] = (yld, cost, price)
    print(f"  [precompute] YCP cache: {len(_ycp_cache)} entries")


def ycp_lookup(crop_id, land_type, season):
    """Fast yield/cost/price lookup. Returns (yield, cost, price) or (0,0,0)."""
    return _ycp_cache.get((crop_id, land_type, season), (0, 0, 0))


def plot_is_two_season(plot):
    """Check if a plot supports two seasons."""
    prefix = plot['prefix']
    return prefix in ['D', 'E', 'F']


def plot_seasons(plot):
    """Return list of seasons for a plot."""
    prefix = plot['prefix']
    if prefix in ['A', 'B', 'C']:
        return ['单季']
    elif prefix == 'D':
        # Water-irrigated: can choose mode
        return ['rice', 'veg_double']  # Two modes
    else:
        # Greenhouses: always two seasons
        return [SEASON1, SEASON2]


def get_allowed_crops(plot_name, season_label):
    """Get set of allowed crops for a plot-season combination."""
    return _compat.get((plot_name, season_label), set())


def compute_plot_profit_single(plot, crop_id, season, scenario=1):
    """
    Compute profit per mu for growing crop_id on plot in given season.
    Scenario 1: excess wasted; Scenario 2: excess sold at 50%.
    Actually returns full profit assuming all produced is sold at full price
    (Lagrangian handles the sales cap).
    """
    ptype = plot['type']
    area = plot['area']
    yld, cost, price = ycp_lookup( crop_id, ptype, season)

    if yld == 0:
        return -1e9  # Infeasible

    production = area * yld
    full_revenue = production * price
    total_cost = area * cost
    profit = full_revenue - total_cost
    return profit, production


# ─── DP for Single-Season Plots ───
# State: (last_crop, years_since_legume)
#   last_crop: 0..41 (0 = none/first year)
#   years_since_legume: 0, 1, 2 (2 = must plant legume this year)

def dp_single_season(plot, lambdas, scenario=1):
    """
    DP for a single-season plot (A/B/C type) over 7 years.
    Returns: (best_value, best_path) where path is list of (crop_id, area) per year.
    """
    area = plot['area']
    ptype = plot['type']
    pname = plot['name']
    init_legume = _legume_init.get(pname, False)
    init_legume_years = 0 if init_legume else 1  # Years since last legume at start

    allowed = get_allowed_crops(pname, '单季')
    allowed_list = sorted(allowed)

    N = len(YEARS)
    # DP[t][last_crop][legume_years] = (max_value, prev_crop, prev_legume_years, crop_chosen)
    # We'll use a simpler approach: finite horizon backward DP

    # State space is manageable: ~42 crops × 3 legume states = 126 states
    # Forward DP with memoization

    from functools import lru_cache

    @lru_cache(maxsize=None)
    def dp(t, last_crop, legume_years):
        """Returns (max_value, [decisions from t to end])"""
        if t >= N:
            return 0.0, []

        best_val = -1e12
        best_decisions = None

        candidates = allowed_list.copy()
        # Can't plant same crop consecutively
        if last_crop > 0 and last_crop in candidates:
            candidates.remove(last_crop)

        # Must plant legume if legume_years == 2
        must_legume = (legume_years >= 2)

        for crop_id in candidates:
            is_legume = crop_id in ALL_LEGUMES
            if must_legume and not is_legume:
                continue

            yld, cost, price = ycp_lookup( crop_id, ptype, '单季')
            if yld == 0:
                continue

            production = area * yld
            total_cost = area * cost

            # Penalize by lambda for overproduction
            sales_cap = _baseline_sales.get(crop_id, 0)
            if scenario == 1:
                # Excess wasted
                sold = min(production, max(0, sales_cap))
            else:
                # Excess at 50% price
                sold = production  # all sold, excess at half price handled below

            if scenario == 1:
                revenue = sold * price
            else:
                if production <= sales_cap:
                    revenue = production * price
                else:
                    revenue = sales_cap * price + (production - sales_cap) * price * 0.5

            profit = revenue - total_cost

            # Lagrangian penalty on production exceeding cap
            lag_penalty = lambdas.get(crop_id, 0.0) * max(0, production - sales_cap)

            next_legume = 0 if is_legume else legume_years + 1
            next_crop = crop_id

            future_val, future_dec = dp(t + 1, next_crop, next_legume)
            total_val = profit - lag_penalty + future_val

            if total_val > best_val:
                best_val = total_val
                best_decisions = [(crop_id, area)] + future_dec

        # Also consider: plant multiple crops (mixed planting)
        # For single-season open plots, mixed planting is less common but allowed
        # We'll handle it as: plant a main crop + optionally a legume if needed
        if not must_legume and len(candidates) >= 2:
            # Try splitting: main crop + legume (for future benefit)
            for main_crop in candidates:
                if main_crop in ALL_LEGUMES:
                    continue
                yld_m, cost_m, price_m = ycp_lookup( main_crop, ptype, '单季')
                if yld_m == 0:
                    continue

                for leg_crop in candidates:
                    if leg_crop not in ALL_LEGUMES:
                        continue
                    yld_l, cost_l, price_l = ycp_lookup( leg_crop, ptype, '单季')
                    if yld_l == 0:
                        continue

                    # Split area: 80% main, 20% legume (minimum area check)
                    area_main = area * 0.8
                    area_leg = area * 0.2

                    if area_leg < MIN_AREA_PER_CROP_OPEN:
                        continue

                    prod_m = area_main * yld_m
                    prod_l = area_leg * yld_l
                    cost_total = area_main * cost_m + area_leg * cost_l

                    cap_m = _baseline_sales.get(main_crop, 0)
                    cap_l = _baseline_sales.get(leg_crop, 0)

                    if scenario == 1:
                        sold_m = min(prod_m, max(0, cap_m))
                        sold_l = min(prod_l, max(0, cap_l))
                        revenue = sold_m * price_m + sold_l * price_l
                    else:
                        if prod_m <= cap_m:
                            rev_m = prod_m * price_m
                        else:
                            rev_m = cap_m * price_m + (prod_m - cap_m) * price_m * 0.5
                        if prod_l <= cap_l:
                            rev_l = prod_l * price_l
                        else:
                            rev_l = cap_l * price_l + (prod_l - cap_l) * price_l * 0.5
                        revenue = rev_m + rev_l

                    profit = revenue - cost_total
                    lag_penalty = (lambdas.get(main_crop, 0.0) * max(0, prod_m - cap_m) +
                                   lambdas.get(leg_crop, 0.0) * max(0, prod_l - cap_l))

                    next_legume = 0  # legume planted
                    # last_crop tracking: we track the main crop for simplicity
                    future_val, future_dec = dp(t + 1, main_crop, next_legume)
                    total_val = profit - lag_penalty + future_val

                    if total_val > best_val:
                        best_val = total_val
                        best_decisions = [(f'{main_crop}+{leg_crop}', area)] + future_dec

        if best_decisions is None:
            # Fallback: plant anything feasible
            for crop_id in candidates:
                if must_legume and crop_id not in ALL_LEGUMES:
                    continue
                yld, cost, price = ycp_lookup( crop_id, ptype, '单季')
                if yld > 0:
                    return dp_from_choice(t, crop_id, area, legume_years, scenario, lambdas, price, yld, cost)

            best_val = 0.0
            best_decisions = [(0, 0)] + ([] if t + 1 >= N else dp(t + 1, 0, min(legume_years + 1, 3))[1])

        return best_val, best_decisions

    def dp_from_choice(t, crop_id, area, legume_years, scenario, lambdas, price, yld, cost):
        production = area * yld
        total_cost = area * cost
        sales_cap = _baseline_sales.get(crop_id, 0)
        if scenario == 1:
            sold = min(production, max(0, sales_cap))
            revenue = sold * price
        else:
            if production <= sales_cap:
                revenue = production * price
            else:
                revenue = sales_cap * price + (production - sales_cap) * price * 0.5
        profit = revenue - total_cost
        lag_penalty = lambdas.get(crop_id, 0.0) * max(0, production - sales_cap)

        is_legume = crop_id in ALL_LEGUMES
        next_legume = 0 if is_legume else legume_years + 1
        future_val, future_dec = dp(t + 1, crop_id, next_legume)
        total_val = profit - lag_penalty + future_val
        return total_val, [(crop_id, area)] + future_dec

    # Get initial legume years
    init_ls = init_legume_years
    best_val, best_path = dp(0, 0, init_ls)
    return best_val, best_path


# ─── DP for Water-Irrigated Land (D type) ───
def dp_irrigated(plot, lambdas, scenario=1):
    """
    DP for water-irrigated land. Two modes:
    - Rice (single season)
    - Two-season vegetables
    """
    area = plot['area']
    ptype = plot['type']
    pname = plot['name']
    init_legume = _legume_init.get(pname, False)
    init_legume_years = 0 if init_legume else 1

    N = len(YEARS)
    from functools import lru_cache

    @lru_cache(maxsize=None)
    def dp(t, last_legume_years):
        """Returns (best_value, list_of_annual_decisions)"""
        if t >= N:
            return 0.0, []

        best_val = -1e12
        best_dec = None

        must_legume = (last_legume_years >= 2)

        # --- Mode 1: Rice (single season) ---
        rice_yld, rice_cost, rice_price = ycp_lookup( RICE, ptype, '单季')
        if rice_yld > 0:
            prod = area * rice_yld
            cap = _baseline_sales.get(RICE, 0)

            if scenario == 1:
                sold = min(prod, max(0, cap))
                revenue = sold * rice_price
            else:
                if prod <= cap:
                    revenue = prod * rice_price
                else:
                    revenue = cap * rice_price + (prod - cap) * rice_price * 0.5

            cost_total = area * rice_cost
            profit = revenue - cost_total
            lag_penalty = lambdas.get(RICE, 0.0) * max(0, prod - cap)

            next_ls = last_legume_years + 1  # Rice is not legume
            if next_ls <= 3:  # Still feasible
                future_val, future_dec = dp(t + 1, next_ls)
                total_val = profit - lag_penalty + future_val
                if total_val > best_val:
                    best_val = total_val
                    best_dec = [{'mode': 'rice', 's1': [(RICE, area)], 's2': []}] + future_dec

        # --- Mode 2: Two-season vegetables ---
        # Season 1: vegetables (17-34 except 35-37)
        # Season 2: exactly one of 35, 36, 37

        s1_crops = VEGETABLE_LEGUMES + VEGETABLES
        s2_crops = WINTER_VEGETABLES

        # For each s2 choice, find best s1
        for s2_crop in s2_crops:
            yld_s2, cost_s2, price_s2 = ycp_lookup( s2_crop, ptype, SEASON2)
            if yld_s2 == 0:
                continue

            prod_s2 = area * yld_s2
            cap_s2 = _baseline_sales.get(s2_crop, 0)

            if scenario == 1:
                sold_s2 = min(prod_s2, max(0, cap_s2))
                rev_s2 = sold_s2 * price_s2
            else:
                if prod_s2 <= cap_s2:
                    rev_s2 = prod_s2 * price_s2
                else:
                    rev_s2 = cap_s2 * price_s2 + (prod_s2 - cap_s2) * price_s2 * 0.5

            # Season 1: pick the best single crop or combination
            # For simplicity, try single crops first
            best_s1_val = -1e12
            best_s1_choice = None

            for s1_crop in s1_crops:
                is_legume = s1_crop in ALL_LEGUMES
                if must_legume and not is_legume:
                    continue

                yld_s1, cost_s1, price_s1 = ycp_lookup( s1_crop, ptype, SEASON1)
                if yld_s1 == 0:
                    continue

                prod_s1 = area * yld_s1
                cap_s1 = _baseline_sales.get(s1_crop, 0)

                if scenario == 1:
                    sold_s1 = min(prod_s1, max(0, cap_s1))
                    rev_s1 = sold_s1 * price_s1
                else:
                    if prod_s1 <= cap_s1:
                        rev_s1 = prod_s1 * price_s1
                    else:
                        rev_s1 = cap_s1 * price_s1 + (prod_s1 - cap_s1) * price_s1 * 0.5

                profit_s1 = rev_s1 - area * cost_s1
                profit_s2 = rev_s2 - area * cost_s2
                total_profit = profit_s1 + profit_s2

                lag_penalty = (lambdas.get(s1_crop, 0.0) * max(0, prod_s1 - cap_s1) +
                               lambdas.get(s2_crop, 0.0) * max(0, prod_s2 - cap_s2))

                next_ls = 0 if is_legume else last_legume_years + 1
                if next_ls <= 3:
                    future_val, future_dec = dp(t + 1, next_ls)
                    total_val = total_profit - lag_penalty + future_val

                    if total_val > best_s1_val:
                        best_s1_val = total_val
                        best_s1_choice = {
                            'mode': 'veg_double',
                            's1': [(s1_crop, area)],
                            's2': [(s2_crop, area)],
                            'next_ls': next_ls,
                            'future_val': future_val,
                            'future_dec': future_dec,
                        }

            # Also try mixed s1: part legume + part other vegetable
            if not must_legume:
                for s1_crop in s1_crops:
                    if s1_crop in ALL_LEGUMES:
                        continue
                    yld_main, cost_main, price_main = ycp_lookup( s1_crop, ptype, SEASON1)
                    if yld_main == 0:
                        continue
                    for leg_crop in VEGETABLE_LEGUMES:
                        yld_leg, cost_leg, price_leg = ycp_lookup( leg_crop, ptype, SEASON1)
                        if yld_leg == 0:
                            continue

                        area_leg = area * 0.25
                        area_main = area * 0.75
                        if area_leg < MIN_AREA_PER_CROP_OPEN:
                            continue

                        prod_main = area_main * yld_main
                        prod_leg = area_leg * yld_leg
                        cap_main = _baseline_sales.get(s1_crop, 0)
                        cap_leg = _baseline_sales.get(leg_crop, 0)

                        if scenario == 1:
                            rev = min(prod_main, cap_main) * price_main + min(prod_leg, cap_leg) * price_leg
                        else:
                            rev = (min(prod_main, cap_main) * price_main +
                                   max(0, prod_main - cap_main) * price_main * 0.5 +
                                   min(prod_leg, cap_leg) * price_leg +
                                   max(0, prod_leg - cap_leg) * price_leg * 0.5)

                        profit_s1 = rev - (area_main * cost_main + area_leg * cost_leg)
                        total_profit = profit_s1 + rev_s2 - area * cost_s2

                        lag_penalty = (lambdas.get(s1_crop, 0.0) * max(0, prod_main - cap_main) +
                                       lambdas.get(leg_crop, 0.0) * max(0, prod_leg - cap_leg) +
                                       lambdas.get(s2_crop, 0.0) * max(0, prod_s2 - cap_s2))

                        next_ls = 0  # legume planted
                        future_val, future_dec = dp(t + 1, next_ls)
                        total_val = total_profit - lag_penalty + future_val

                        if total_val > best_s1_val:
                            best_s1_val = total_val
                            best_s1_choice = {
                                'mode': 'veg_double',
                                's1': [(s1_crop, area_main), (leg_crop, area_leg)],
                                's2': [(s2_crop, area)],
                                'next_ls': next_ls,
                                'future_val': future_val,
                                'future_dec': future_dec,
                            }

            if best_s1_choice is not None and best_s1_val > best_val:
                best_val = best_s1_val
                dec = {
                    'mode': best_s1_choice['mode'],
                    's1': best_s1_choice['s1'],
                    's2': best_s1_choice['s2'],
                }
                best_dec = [dec] + best_s1_choice['future_dec']

        if best_dec is None:
            # Fallback: rice if feasible
            if not must_legume:
                future_val, future_dec = dp(t + 1, min(last_legume_years + 1, 3))
                best_val = future_val
                best_dec = [{'mode': 'rice', 's1': [(RICE, area)], 's2': []}] + future_dec

        return best_val, best_dec

    val, path = dp(0, init_legume_years)
    return val, path


# ─── DP for Greenhouses (E and F types) ───
def dp_greenhouse(plot, lambdas, scenario=1):
    """
    DP for greenhouse (regular E or smart F).
    E: s1 vegetables, s2 mushrooms (mandatory)
    F: s1 vegetables, s2 vegetables

    Optimized: precompute profit matrices to avoid repeated yield/cost/price
    lookups and revenue calculations in the DP inner loops.
    """
    area = plot['area']
    ptype = plot['type']
    pname = plot['name']
    prefix = plot['prefix']
    init_legume = _legume_init.get(pname, False)
    init_legume_years = 0 if init_legume else 1

    is_smart = (prefix == 'F')
    N = len(YEARS)

    # ─── Precompute: profit components for every valid (s1, s2) pair ───
    # This moves ~90% of the inner-loop work OUTSIDE the recursive DP.
    # single_matrix[s1][s2] = (profit, lag_overprod, is_legume)
    # mixed_matrix[(s1a,s1b)][s2] = (profit, lag_overprod, is_legume)

    all_s1 = VEGETABLE_LEGUMES + VEGETABLES
    if is_smart:
        # For smart greenhouses, rank s2 crops by profit potential and keep top 10
        s2_scores = []
        for s2 in VEGETABLE_LEGUMES + VEGETABLES:
            y, c, p = ycp_lookup(s2, ptype, SEASON2)
            s2_scores.append((s2, p * y - c if y > 0 else -1e9))
        s2_scores.sort(key=lambda x: -x[1])
        all_s2 = [s for s, _ in s2_scores[:10]]  # Top 10 by unit profit
    else:
        all_s2 = MUSHROOMS

    # Helper: compute revenue
    def _rev(prod, cap, price):
        if scenario == 1:
            return min(prod, max(0, cap)) * price
        else:
            return (cap * price + (prod - cap) * price * 0.5) if prod > cap else prod * price

    # --- Single-crop matrix ---
    single_profit = {}   # (s1, s2) -> (profit, lag_base_s1, lag_base_s2, is_legume, prod_s1, prod_s2, cap_s1, cap_s2)
    for s1 in all_s1:
        y1, c1, p1 = ycp_lookup(s1, ptype, SEASON1)
        if y1 == 0:
            continue
        prod1 = area * y1
        cap1 = _baseline_sales.get(s1, 0)
        for s2 in all_s2:
            y2, c2, p2 = ycp_lookup(s2, ptype, SEASON2)
            if y2 == 0:
                continue
            prod2 = area * y2
            cap2 = _baseline_sales.get(s2, 0)
            profit = _rev(prod1, cap1, p1) + _rev(prod2, cap2, p2) - area * (c1 + c2)
            lag_base_s1 = max(0, prod1 - cap1)
            lag_base_s2 = max(0, prod2 - cap2)
            is_leg = (s1 in ALL_LEGUMES) or (s2 in ALL_LEGUMES)
            single_profit[(s1, s2)] = (profit, lag_base_s1, lag_base_s2, is_leg)

    # --- Mixed-s1 matrix (split area between s1a and s1b) ---
    mixed_profit = {}  # (s1a, s1b, s2) -> (profit, lag_base_a, lag_base_b, lag_base_s2, is_legume)
    if area >= 0.3:
        for i, s1a in enumerate(all_s1[:8]):    # Limit to 8 for speed
            y_a, c_a, p_a = ycp_lookup(s1a, ptype, SEASON1)
            if y_a == 0: continue
            for s1b in all_s1[:8]:
                if s1a >= s1b: continue
                y_b, c_b, p_b = ycp_lookup(s1b, ptype, SEASON1)
                if y_b == 0: continue
                half = area / 2
                prod_a, prod_b = half * y_a, half * y_b
                cap_a = _baseline_sales.get(s1a, 0)
                cap_b = _baseline_sales.get(s1b, 0)
                has_leg = (s1a in ALL_LEGUMES) or (s1b in ALL_LEGUMES)

                for s2 in all_s2[:5]:   # Limit to 5
                    y2, c2, p2 = ycp_lookup(s2, ptype, SEASON2)
                    if y2 == 0: continue
                    prod2 = area * y2
                    cap2 = _baseline_sales.get(s2, 0)
                    profit = (_rev(prod_a, cap_a, p_a) + _rev(prod_b, cap_b, p_b) +
                              _rev(prod2, cap2, p2) - half * (c_a + c_b) - area * c2)
                    lag_a = max(0, prod_a - cap_a)
                    lag_b = max(0, prod_b - cap_b)
                    lag_2 = max(0, prod2 - cap2)
                    is_leg = has_leg or (s2 in ALL_LEGUMES)
                    mixed_profit[(s1a, s1b, s2)] = (profit, lag_a, lag_b, lag_2, is_leg)

    from functools import lru_cache

    @lru_cache(maxsize=None)
    def dp(t, last_s1_crop, last_s2_crop, legume_years):
        if t >= N:
            return 0.0, []

        best_val = -1e12
        best_dec = None
        must_legume = (legume_years >= 2)

        # --- Single-crop combinations ---
        for (s1, s2), (profit, lag_s1, lag_s2, is_leg) in single_profit.items():
            if s1 == last_s1_crop or s2 == last_s2_crop:
                continue
            if must_legume and not is_leg:
                continue
            next_ls = 0 if is_leg else legume_years + 1
            if next_ls > 3:
                continue

            lag_pen = lambdas.get(s1, 0.0) * lag_s1 + lambdas.get(s2, 0.0) * lag_s2
            future_val, future_dec = dp(t + 1, s1, s2, next_ls)
            total_val = profit - lag_pen + future_val

            if total_val > best_val:
                best_val = total_val
                best_dec = [{'mode': 'double', 's1': [(s1, area)], 's2': [(s2, area)]}] + future_dec

        # --- Mixed-s1 combinations ---
        for (s1a, s1b, s2), (profit, lag_a, lag_b, lag_2, is_leg) in mixed_profit.items():
            if last_s1_crop in (s1a, s1b) or last_s2_crop == s2:
                continue
            if must_legume and not is_leg:
                continue
            next_ls = 0 if is_leg else legume_years + 1
            if next_ls > 3:
                continue

            lag_pen = (lambdas.get(s1a, 0.0) * lag_a + lambdas.get(s1b, 0.0) * lag_b +
                       lambdas.get(s2, 0.0) * lag_2)
            half = area / 2
            future_val, future_dec = dp(t + 1, s1a, s2, next_ls)
            total_val = profit - lag_pen + future_val

            if total_val > best_val:
                best_val = total_val
                best_dec = [{'mode': 'double', 's1': [(s1a, half), (s1b, half)],
                             's2': [(s2, area)]}] + future_dec

        if best_dec is None:
            future_val, future_dec = dp(t + 1, 0, 0, min(legume_years + 1, 3))
            best_val = future_val
            best_dec = [{'mode': 'empty', 's1': [], 's2': []}] + future_dec

        return best_val, best_dec

    val, path = dp(0, 0, 0, init_legume_years)
    return val, path


# ─── Master Solver: Lagrangian Relaxation ───

def solve_lagrangian(scenario=1, max_iter=LAGRANGE_MAX_ITER):
    """
    Lagrangian relaxation solver.
    Relaxes: total_production[crop] <= expected_sales[crop]
    """
    print(f"\n{'='*60}")
    print(f"Problem 1 - Scenario {scenario}: ", end="")
    if scenario == 1:
        print("Excess wasted")
    else:
        print("Excess sold at 50% discount")
    print(f"{'='*60}")

    # Initialize Lagrange multipliers
    lambdas = {cid: 0.0 for cid in ALL_CROPS}

    best_feasible_val = -1e12
    best_feasible_solution = None
    best_lb = -1e12  # Best lower bound (from feasible solutions)
    ub_history = []
    lb_history = []

    rho = SUBLINEAR_RHO_START
    import time
    t_start = time.time()

    for iteration in range(max_iter):
        t_iter = time.time()
        # ─── Solve each plot independently ───
        plot_solutions = {}
        total_dual = 0.0
        total_production = {cid: 0.0 for cid in ALL_CROPS}
        total_profit_raw = 0.0  # Profit before Lagrangian correction

        for plot in _plots:
            pname = plot['name']
            prefix = plot['prefix']

            # Choose DP based on plot type
            if prefix in ['A', 'B', 'C']:
                val, sol = dp_single_season(plot, lambdas, scenario)
            elif prefix == 'D':
                val, sol = dp_irrigated(plot, lambdas, scenario)
            else:  # E, F
                val, sol = dp_greenhouse(plot, lambdas, scenario)

            plot_solutions[pname] = sol
            total_dual += val

            # Compute actual production from this solution
            for year_idx, annual in enumerate(sol):
                if isinstance(annual, dict):
                    for season_crops in [annual.get('s1', []), annual.get('s2', [])]:
                        for crop_entry in season_crops:
                            if isinstance(crop_entry[0], str) and '+' in crop_entry[0]:
                                # Mixed planting encoded as string
                                parts = crop_entry[0].split('+')
                                for i, c in enumerate(parts):
                                    try:
                                        cid = int(c)
                                        total_production[cid] += (crop_entry[1] / len(parts) *
                                                                  ycp_lookup( cid, plot['type'],
                                                                                       '单季')[0])
                                    except:
                                        pass
                            elif isinstance(crop_entry[0], int) and crop_entry[0] > 0:
                                cid = crop_entry[0]
                                season_key = '单季'
                                if 's1' in annual and crop_entry in annual.get('s1', []):
                                    season_key = SEASON1
                                elif 's2' in annual and crop_entry in annual.get('s2', []):
                                    season_key = SEASON2

                                yld, _, _ = ycp_lookup( cid, plot['type'], season_key)
                                total_production[cid] += crop_entry[1] * yld

        # ─── Compute upper bound (dual value + penalty correction) ───
        penalty_sum = sum(lambdas[cid] * _baseline_sales.get(cid, 0) for cid in ALL_CROPS)
        upper_bound = total_dual + penalty_sum
        ub_history.append(upper_bound)

        # ─── Check feasibility and compute true profit ───
        feasible = True
        violations = {}
        for cid in ALL_CROPS:
            cap = _baseline_sales.get(cid, 0)
            if cap > 0 and total_production[cid] > cap * 1.05:  # 5% tolerance
                violations[cid] = total_production[cid] - cap
            elif cap == 0 and total_production[cid] > 100:  # Some tolerance for zero-cap crops
                pass

        # Compute real profit (accounting for sales caps)
        # Simplified: compute from raw production
        real_profit = 0.0
        for cid in ALL_CROPS:
            prod = total_production[cid]
            cap = _baseline_sales.get(cid, 0)
            # We need average price for this crop
            avg_price = 0.0
            count = 0
            for plot in _plots:
                for season in ['单季', SEASON1, SEASON2]:
                    _, _, p = ycp_lookup( cid, plot['type'], season)
                    if p > 0:
                        avg_price += p
                        count += 1
            avg_price = avg_price / count if count > 0 else 0

            if scenario == 1:
                sold = min(prod, cap)
                real_profit += sold * avg_price
            else:
                if prod <= cap:
                    real_profit += prod * avg_price
                else:
                    real_profit += cap * avg_price + (prod - cap) * avg_price * 0.5

        lb_history.append(real_profit)

        if real_profit > best_lb:
            best_lb = real_profit
            best_feasible_val = real_profit
            best_feasible_solution = (plot_solutions, total_production, lambdas.copy())

        # ─── Update Lagrange multipliers (subgradient) ───
        subgrad_norm_sq = 0.0
        for cid in ALL_CROPS:
            cap = _baseline_sales.get(cid, 0)
            violation = max(0, total_production[cid] - cap)
            subgrad_norm_sq += violation ** 2

        if subgrad_norm_sq > 0:
            step_size = rho * (upper_bound - best_lb) / subgrad_norm_sq
            for cid in ALL_CROPS:
                cap = _baseline_sales.get(cid, 0)
                violation = max(0, total_production[cid] - cap)
                lambdas[cid] = max(0.0, lambdas[cid] + step_size * violation)

        # Update rho (halve every 50 iterations)
        if (iteration + 1) % 50 == 0:
            rho /= 2.0

        # Convergence check
        gap = upper_bound - best_lb
        if iteration % 10 == 0 or iteration < 3:
            t_elapsed = time.time() - t_start
            print(f"  Iter {iteration:3d}: UB={upper_bound:.0f}, LB={best_lb:.0f}, "
                  f"gap={gap:.0f} ({100*gap/abs(best_lb) if best_lb > 0 else 0:.1f}%), "
                  f"viols={sum(1 for v in violations.values() if v > 0)}, "
                  f"t={time.time()-t_iter:.1f}s")

        if gap < LAGRANGE_TOL * abs(best_lb) and iteration > 10:
            print(f"  Converged at iter {iteration}")
            break

    # ─── Final feasible solution recovery ───
    print(f"\n  Final: UB={ub_history[-1]:.0f}, Best LB={best_lb:.0f}")
    print(f"  Gap: {ub_history[-1] - best_lb:.0f}")

    return best_feasible_solution, ub_history, lb_history


def generate_result_table(solution, scenario=1):
    """
    Convert solution to result table format for Excel output.
    solution: (plot_solutions, total_production, lambdas)
    Returns a nested dict: year -> season -> plot -> {crop_id: area}
    """
    plot_solutions, total_production, _ = solution
    result = {year: {} for year in YEARS}

    for pname, annual_decisions in plot_solutions.items():
        for yr_idx, year in enumerate(YEARS):
            if year not in result:
                result[year] = {}
            if '第一季' not in result[year]:
                result[year]['第一季'] = {}
                result[year]['第二季'] = {}

            if yr_idx < len(annual_decisions):
                dec = annual_decisions[yr_idx]
                if isinstance(dec, dict):
                    # Two-season decision
                    for crop_id, area in dec.get('s1', []):
                        cid = crop_id if isinstance(crop_id, int) else 0
                        if cid > 0:
                            if pname not in result[year]['第一季']:
                                result[year]['第一季'][pname] = {}
                            result[year]['第一季'][pname][cid] = area

                    for crop_id, area in dec.get('s2', []):
                        cid = crop_id if isinstance(crop_id, int) else 0
                        if cid > 0:
                            if pname not in result[year]['第二季']:
                                result[year]['第二季'][pname] = {}
                            result[year]['第二季'][pname][cid] = area

                    # Handle "rice" mode - it's single season
                    if dec.get('mode') == 'rice':
                        for crop_id, area in dec.get('s1', []):
                            cid = crop_id if isinstance(crop_id, int) else 0
                            if cid > 0:
                                if pname not in result[year]['第一季']:
                                    result[year]['第一季'][pname] = {}
                                result[year]['第一季'][pname][cid] = area
                elif isinstance(dec, (list, tuple)):
                    # Single season: dec is [(crop_id, area)]
                    for item in dec:
                        if isinstance(item, (list, tuple)) and len(item) == 2:
                            crop_id, area = item
                            if isinstance(crop_id, str) and '+' in crop_id:
                                # Mixed - split into components
                                parts = crop_id.split('+')
                                half_area = area / len(parts)
                                for p in parts:
                                    try:
                                        cid = int(p)
                                        if pname not in result[year]['第一季']:
                                            result[year]['第一季'][pname] = {}
                                        result[year]['第一季'][pname][cid] = result[year]['第一季'][pname].get(cid, 0) + half_area
                                    except:
                                        pass
                            elif isinstance(crop_id, int) and crop_id > 0:
                                if pname not in result[year]['第一季']:
                                    result[year]['第一季'][pname] = {}
                                result[year]['第一季'][pname][cid] = result[year]['第一季'][pname].get(crop_id, 0) + area

    return result, total_production


def export_to_excel(result_table, output_path):
    """Export result table to Excel file matching the template format."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    for year in YEARS:
        if year == YEARS[0]:
            ws = wb.active
            ws.title = str(year)
        else:
            ws = wb.create_sheet(str(year))

        # Header row
        ws.cell(row=1, column=1, value='')
        ws.cell(row=1, column=2, value='地块名')
        for j, cid in enumerate(ALL_CROPS):
            ws.cell(row=1, column=3 + j, value=CROP_NAMES.get(cid, f'Crop{cid}'))

        # Merge header for season labels
        ws.cell(row=2, column=1, value='第一季')
        ws.cell(row=3, column=1, value='第二季')

        # Fill with zeros initially
        row_offset = {'第一季': 2, '第二季': 3}
        all_plot_names = sorted([p['name'] for p in _plots],
                                key=lambda x: (x[0], int(x[1:])))

        for season in ['第一季', '第二季']:
            r = row_offset[season]
            for i, pname in enumerate(all_plot_names):
                ws.cell(row=r + i, column=2, value=pname)
                for j, cid in enumerate(ALL_CROPS):
                    ws.cell(row=r + i, column=3 + j, value=0.0)

        # Fill data
        if year in result_table:
            for season in ['第一季', '第二季']:
                if season in result_table[year]:
                    r = row_offset[season]
                    for i, pname in enumerate(all_plot_names):
                        if pname in result_table[year][season]:
                            for cid, area in result_table[year][season][pname].items():
                                j = ALL_CROPS.index(cid)
                                ws.cell(row=r + i, column=3 + j, value=round(area, 2))

    wb.save(output_path)
    print(f"  Result exported to {output_path}")


def run_problem1(data):
    """Main entry point for Problem 1."""
    set_global_data(data)

    results = {}
    for scenario in [1, 2]:
        sol, ub_hist, lb_hist = solve_lagrangian(scenario=scenario, max_iter=100)
        result_table, total_prod = generate_result_table(sol, scenario)

        # Export to Excel
        output_name = RESULT1_1 if scenario == 1 else RESULT1_2
        export_to_excel(result_table, output_name)

        results[scenario] = {
            'solution': sol,
            'result_table': result_table,
            'total_production': total_prod,
            'ub_history': ub_hist,
            'lb_history': lb_hist,
        }

    return results


if __name__ == '__main__':
    data = preprocess_all()
    results = run_problem1(data)
