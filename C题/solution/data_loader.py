"""
Data loading and preprocessing for 2024 CUMCM Problem C
"""

import numpy as np
import pandas as pd
import openpyxl
from config import *

def load_plots():
    """Load plot information from 附件1."""
    wb = openpyxl.load_workbook(ATTACHMENT1)
    ws = wb['乡村的现有耕地']

    plots = []
    for row in ws.iter_rows(min_row=2, max_row=55, values_only=True):
        name, ptype, area, _ = row
        if name is None:
            continue
        plots.append({
            'name': str(name).strip(),
            'type': str(ptype).strip() if ptype else '',
            'area': float(area),
            'prefix': str(name)[0],
        })
    return plots


def load_crop_info():
    """Load crop compatibility information from 附件1."""
    wb = openpyxl.load_workbook(ATTACHMENT1)
    ws = wb['乡村种植的农作物']

    crops = {}
    for row in ws.iter_rows(min_row=2, max_row=42, values_only=True):
        crop_id, name, ctype, land_desc, _ = row
        if crop_id is None:
            continue
        cid = int(crop_id)
        crops[cid] = {
            'name': str(name).strip(),
            'type': str(ctype).strip() if ctype else '',
            'land_desc': str(land_desc).strip() if land_desc else '',
        }
    return crops


def load_2023_planting():
    """Load 2023 actual planting data from 附件2."""
    wb = openpyxl.load_workbook(ATTACHMENT2)
    ws = wb['2023年的农作物种植情况']

    records = []
    current_plot = None
    for row in ws.iter_rows(min_row=2, max_row=88, values_only=True):
        plot, crop_id, crop_name, ctype, area, season = row
        if plot is not None and str(plot).strip():
            current_plot = str(plot).strip()
        records.append({
            'plot': current_plot,
            'crop_id': int(crop_id) if crop_id else None,
            'crop_name': str(crop_name).strip() if crop_name else '',
            'crop_type': str(ctype).strip() if ctype else '',
            'area': float(area) if area else 0,
            'season': str(season).strip() if season else '',
        })
    return records


def load_yield_cost_price():
    """Load yield, cost, and price data from 附件2."""
    wb = openpyxl.load_workbook(ATTACHMENT2)
    ws = wb['2023年统计的相关数据']

    data = {}
    for row in ws.iter_rows(min_row=2, max_row=108, values_only=True):
        seq, crop_id, crop_name, land_type, season, yield_per_mu, cost_per_mu, price_range = row[:8]
        if crop_id is None:
            continue

        cid = int(crop_id)
        lt = str(land_type).strip() if land_type else ''
        s = str(season).strip() if season else ''

        # Parse price range
        if price_range and '-' in str(price_range):
            parts = str(price_range).split('-')
            price_low = float(parts[0])
            price_high = float(parts[1])
            price_mid = (price_low + price_high) / 2
        else:
            price_low = price_high = price_mid = 0

        key = (cid, lt, s)
        data[key] = {
            'yield': float(yield_per_mu) if yield_per_mu else 0,
            'cost': float(cost_per_mu) if cost_per_mu else 0,
            'price_low': price_low,
            'price_high': price_high,
            'price_mid': price_mid,
        }

    # Note: 智慧大棚第一季 uses same data as 普通大棚第一季
    # Copy relevant entries
    for cid in VEGETABLE_LEGUMES + VEGETABLES:
        src_key = (cid, '普通大棚', SEASON1)
        dst_key = (cid, '智慧大棚', SEASON1)
        if src_key in data and dst_key not in data:
            data[dst_key] = data[src_key].copy()

    return data


def build_compatibility_matrix(plots, crop_info):
    """
    Build compatibility matrix: which crop can be planted on which plot in which season.
    Returns dict: (plot_name, season) -> set of allowed crop_ids
    """
    compat = {}

    for plot in plots:
        pname = plot['name']
        ptype = plot['type']
        prefix = plot['prefix']

        if prefix in ['A', 'B', 'C']:
            # 平旱地/梯田/山坡地: single season, grain crops only (no rice)
            allowed = GRAIN_LEGUMES + GRAIN_CROPS
            compat[(pname, '单季')] = set(allowed)
            compat[(pname, SEASON1)] = set()
            compat[(pname, SEASON2)] = set()

        elif prefix == 'D':
            # 水浇地: single season rice OR two-season vegetables
            compat[(pname, '单季')] = {RICE}
            compat[(pname, SEASON1)] = set(VEGETABLE_LEGUMES + VEGETABLES)
            compat[(pname, SEASON2)] = set(WINTER_VEGETABLES)  # Only one of 35,36,37

        elif prefix == 'E':
            # 普通大棚: season1 vegetables, season2 mushrooms
            compat[(pname, '单季')] = set()
            compat[(pname, SEASON1)] = set(VEGETABLE_LEGUMES + VEGETABLES)
            compat[(pname, SEASON2)] = set(MUSHROOMS)

        elif prefix == 'F':
            # 智慧大棚: both seasons vegetables (no winter veg)
            compat[(pname, '单季')] = set()
            compat[(pname, SEASON1)] = set(VEGETABLE_LEGUMES + VEGETABLES)
            compat[(pname, SEASON2)] = set(VEGETABLE_LEGUMES + VEGETABLES)

    return compat


def compute_2023_baseline(planting_2023, yield_data, plots):
    """
    Compute 2023 actual production for each crop -> baseline expected sales.
    Also compute profit for each crop-plot combination.
    """
    # Map plots to their types
    plot_type_map = {p['name']: p['type'] for p in plots}

    # Compute total production by crop
    production = {cid: 0.0 for cid in ALL_CROPS}
    revenue = {cid: 0.0 for cid in ALL_CROPS}
    total_cost = {cid: 0.0 for cid in ALL_CROPS}

    for rec in planting_2023:
        cid = rec['crop_id']
        area = rec['area']
        plot_name = rec['plot']
        season = rec['season']
        ptype = plot_type_map.get(plot_name, '')

        if cid is None or area == 0:
            continue

        # Map season name
        if season == '单季':
            s = '单季'
        elif season == '第一季':
            s = SEASON1
        elif season == '第二季':
            s = SEASON2
        else:
            continue

        key = (cid, ptype, s)
        if key in yield_data:
            yld = yield_data[key]['yield']
            cost = yield_data[key]['cost']
            price = yield_data[key]['price_mid']
            prod = area * yld
            production[cid] += prod
            revenue[cid] += prod * price
            total_cost[cid] += area * cost

    profit = {cid: revenue[cid] - total_cost[cid] for cid in ALL_CROPS}

    return {
        'production': production,    # 斤
        'revenue': revenue,          # 元
        'cost': total_cost,          # 元
        'profit': profit,            # 元
    }


def compute_legume_status_2023(planting_2023):
    """
    Determine which plots had legumes in 2023.
    Returns dict: plot_name -> True/False
    """
    status = {}
    for rec in planting_2023:
        plot = rec['plot']
        cid = rec['crop_id']
        if cid in ALL_LEGUMES:
            status[plot] = True

    # Mark all plots, default False
    all_plots_seen = set()
    for rec in planting_2023:
        all_plots_seen.add(rec['plot'])
    for p in all_plots_seen:
        if p not in status:
            status[p] = False

    return status


def get_yield_cost_price(yield_data, crop_id, land_type, season):
    """Get (yield, cost, price_mid) for a crop on specific land/season."""
    key = (crop_id, land_type, season)
    if key in yield_data:
        d = yield_data[key]
        return d['yield'], d['cost'], d['price_mid']

    # Try fallback: for 智慧大棚第一季, use 普通大棚第一季 data
    if land_type == '智慧大棚' and season == SEASON1:
        key2 = (crop_id, '普通大棚', SEASON1)
        if key2 in yield_data:
            d = yield_data[key2]
            return d['yield'], d['cost'], d['price_mid']

    return 0, 0, 0


def get_price_range(yield_data, crop_id, land_type, season):
    """Get (price_low, price_high) for a crop."""
    key = (crop_id, land_type, season)
    if key in yield_data:
        d = yield_data[key]
        return d['price_low'], d['price_high']

    if land_type == '智慧大棚' and season == SEASON1:
        key2 = (crop_id, '普通大棚', SEASON1)
        if key2 in yield_data:
            d = yield_data[key2]
            return d['price_low'], d['price_high']

    return 0, 0


def compute_plot_area_summary(plots):
    """Summarize land areas by type."""
    summary = {}
    for p in plots:
        ptype = p['type']
        if ptype not in summary:
            summary[ptype] = {'count': 0, 'total_area': 0.0, 'plots': []}
        summary[ptype]['count'] += 1
        summary[ptype]['total_area'] += p['area']
        summary[ptype]['plots'].append(p['name'])
    return summary


def preprocess_all():
    """Main preprocessing function. Returns all processed data structures."""
    print("=" * 60)
    print("Data Preprocessing")
    print("=" * 60)

    # Load raw data
    plots = load_plots()
    crop_info = load_crop_info()
    planting_2023 = load_2023_planting()
    yield_data = load_yield_cost_price()

    # Build compatibility matrix
    compat = build_compatibility_matrix(plots, crop_info)

    # Compute 2023 baseline sales
    baseline = compute_2023_baseline(planting_2023, yield_data, plots)

    # Legume status
    legume_status = compute_legume_status_2023(planting_2023)

    # Plot area summary
    area_summary = compute_plot_area_summary(plots)

    # Print summary
    print(f"\nPlots loaded: {len(plots)}")
    print(f"Land types: {list(area_summary.keys())}")
    for lt, info in area_summary.items():
        print(f"  {lt}: {info['count']} plots, {info['total_area']:.1f} mu")

    print(f"\n2023 Baseline Production (top 10 crops):")
    sorted_prod = sorted(baseline['production'].items(), key=lambda x: x[1], reverse=True)
    for cid, prod in sorted_prod[:10]:
        name = CROP_NAMES.get(cid, f'Crop{cid}')
        print(f"  {name}: {prod:.0f} jin, profit={baseline['profit'][cid]:.0f} yuan")

    total_profit = sum(baseline['profit'].values())
    print(f"\n2023 Total Profit: {total_profit:.0f} yuan")

    legume_count = sum(1 for v in legume_status.values() if v)
    print(f"Plots with legumes in 2023: {legume_count}/{len(legume_status)}")

    return {
        'plots': plots,
        'crop_info': crop_info,
        'planting_2023': planting_2023,
        'yield_data': yield_data,
        'compat': compat,
        'baseline': baseline,
        'legume_status': legume_status,
        'area_summary': area_summary,
    }


if __name__ == '__main__':
    data = preprocess_all()
