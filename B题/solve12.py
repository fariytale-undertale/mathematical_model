# -*- coding: utf-8 -*-
"""
2023 国赛 B 题 多波束测线问题
问题 1、2 求解：统一"竖直扇形平面 × 坡面求交"框架
"""
import numpy as np
import openpyxl

# ---------- 统一覆盖宽度模型 ----------
# 换能器在原点(海面), 海底中心 (0,0,-D)
# 坡面: z = -D - tan(alpha_eff) * u   (u 为垂直于测线的水平坐标, 向坡下为正)
# 波束半开角 theta/2, 左右波束与竖直向夹角 ±theta/2
# 覆盖半宽:
#   坡下侧(深) w_deep   = D*sin(h) / (cos(h) - sin(h)*tan(alpha_eff))
#   坡上侧(浅) w_shallow= D*sin(h) / (cos(h) + sin(h)*tan(alpha_eff))
def coverage_width(D, alpha_deg, theta_deg=120.0, beta_deg=90.0):
    """返回 (W, w_shallow, w_deep)
    D: 水深(m)  alpha_deg: 坡度(°)  theta_deg: 开角  beta_deg: 测线与坡面法向水平投影夹角(°)
    视坡度 tan(alpha_eff)=tan(alpha)*|sin(beta)|
    """
    h = np.radians(theta_deg / 2.0)
    alpha = np.radians(alpha_deg)
    beta = np.radians(beta_deg)
    tan_ae = np.tan(alpha) * abs(np.sin(beta))   # 视坡度正切
    sh, ch = np.sin(h), np.cos(h)
    w_shallow = D * sh / (ch + sh * tan_ae)      # 坡上侧(浅)半宽
    w_deep    = D * sh / (ch - sh * tan_ae)      # 坡下侧(深)半宽
    return w_shallow + w_deep, w_shallow, w_deep


# ================= 问题 1 =================
print("=" * 60)
print("问题 1: theta=120°, alpha=1.5°, D0=70 m, 测线沿等高线(β=90°)")
print("=" * 60)
theta1, alpha1, D0_1 = 120.0, 1.5, 70.0
xs = np.arange(-800, 801, 200)                 # 测线距中心点距离 (向东为正=坡下)
tan_a = np.tan(np.radians(alpha1))
D1 = D0_1 + xs * tan_a                          # 各测线水深
W1, ws1, wd1 = coverage_width(D1, alpha1, theta1, 90.0)
d1 = 200.0                                      # 相邻测线间距
# 重叠率: 相邻两条测线覆盖宽度平均, η = 1 - d / ((W_i+W_{i-1})/2)
eta1 = 1.0 - 2.0 * d1 / (W1[1:] + W1[:-1])

print(f"{'x/m':>8} {'D/m':>8} {'W/m':>9} {'w浅/m':>8} {'w深/m':>8} {'η/%':>8}")
for i in range(len(xs)):
    s = "  —  " if i == 0 else f"{eta1[i-1]*100:7.2f}"
    print(f"{xs[i]:8.0f} {D1[i]:8.2f} {W1[i]:9.2f} {ws1[i]:8.2f} {wd1[i]:8.2f} {s}")

# 写 result1.xlsx (按模板结构)
wb = openpyxl.load_workbook("result1.xlsx")
ws = wb["Sheet1"]
for j, v in enumerate(D1, start=2):   # 第2行 海水深度
    ws.cell(row=2, column=j, value=round(float(v), 2))
for j, v in enumerate(W1, start=2):   # 第3行 覆盖宽度
    ws.cell(row=3, column=j, value=round(float(v), 2))
ws.cell(row=4, column=2, value="—")   # 第4行 重叠率
for j, v in enumerate(eta1, start=3):
    ws.cell(row=4, column=j, value=round(float(v) * 100, 2))
wb.save("result1.xlsx")
print("已保存 result1.xlsx")

# ================= 问题 2 =================
print()
print("=" * 60)
print("问题 2: theta=120°, alpha=1.5°, D0=120 m, 测线方向夹角 β, 船距中心 s")
print("=" * 60)
theta2, alpha2, D0_2 = 120.0, 1.5, 120.0
betas = np.arange(0, 360, 45)                  # 测线方向夹角
ss = np.array([0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1])  # 距中心点距离(海里)
NM = 1852.0
tan_a2 = np.tan(np.radians(alpha2))

W2 = np.zeros((len(betas), len(ss)))
for i, b in enumerate(betas):
    for j, s in enumerate(ss):
        # 水深: 船沿测线方向移动, 投影到上坡方向变浅
        D = D0_2 - s * NM * np.cos(np.radians(b)) * tan_a2
        W2[i, j], _, _ = coverage_width(D, alpha2, theta2, b)

print(f"{'β/°':>6}" + "".join(f"{s:>9.2f}" for s in ss))
print(f"{'':>6}" + "".join(f"{'   (距离NM)':>9}" for _ in range(1)))
for i, b in enumerate(betas):
    print(f"{b:6.0f}" + "".join(f"{W2[i,j]:9.2f}" for j in range(len(ss))))

# 写 result2.xlsx (按模板结构)
wb2 = openpyxl.load_workbook("result2.xlsx")
ws2 = wb2["Sheet1"]
# 距离在第2行 C:J (列3..10); 角度在 B 列 行3..10 (β=0..315)
for j, v in enumerate(ss, start=3):
    ws2.cell(row=2, column=j, value=float(v))
for i, b in enumerate(betas):
    row = 3 + i
    ws2.cell(row=row, column=2, value=float(b))
    for j in range(len(ss)):
        ws2.cell(row=row, column=3 + j, value=round(float(W2[i, j]), 2))
wb2.save("result2.xlsx")
print("已保存 result2.xlsx")
