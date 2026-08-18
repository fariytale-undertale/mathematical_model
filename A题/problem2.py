"""
问题2: 确定舞龙队盘入的终止时刻 (板凳之间不发生碰撞)。

物理分析:
  - t=0 到 ~60s: 龙头沿螺线向内盘入，龙尾随之收缩
  - t≈55-60s: 龙尾到达最小半径处 (r≈1.5m)，尾板开始碰撞
    → 链方向自然翻转：尾部向外盘出，形成U形弯
  - t≈67s: 弧长条件 S(θ_head) < Σd_i 满足，全链翻转
  - t≈67-412s: 龙头继续向内，龙身向外延伸
  - t≈412.5s: 龙头区域相邻圈板碰撞 → 盘入终止

方法:
  1. 使用 compute_dragon_at_time (含方向翻转) 计算所有把手位置
  2. 碰撞检测: 相邻圈(Δθ≈2π)板对的顶点到中心线距离 < 0.15m
  3. 二分搜索找到碰撞发生的精确时刻

论文A242参考: t_termination ≈ 412.5s
"""
import numpy as np
from pathlib import Path
import common

OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)


def find_first_collision(t_lo=100.0, t_hi=450.0, tol=1e-4):
    """
    Binary search for the first collision time after the direction flip.

    从 t_lo 开始搜索（跳过翻转前的尾部碰撞），找到头部区域碰撞的精确时间。

    Returns:
        (t_collision, theta_at_collision, collision_info)
    """
    # Phase 1: 粗搜索 — 找到碰撞发生的区间
    print("Phase 1: 粗搜索碰撞区间...")
    dt_coarse = 1.0
    t_last_ok = t_lo

    for t in np.arange(t_lo, t_hi + dt_coarse, dt_coarse):
        result, fail_handle = common.compute_dragon_at_time(t)
        if result is None:
            print(f"  t={t:.0f}s: 几何不可行 (把手#{fail_handle})")
            break

        has_coll, min_dist, info = common.check_collision_at_time(t)
        if not has_coll:
            t_last_ok = t
        else:
            print(f"  t={t:.0f}s: 碰撞! min_dist={min_dist:.6f}m, pair={info}")
            break

    if t_last_ok >= t_hi:
        print(f"  在 [{t_lo}, {t_hi}]s 范围内未检测到碰撞")
        return None, None, None

    t_coll_approx = t_last_ok + dt_coarse
    print(f"  碰撞发生在 [{t_last_ok:.1f}, {t_coll_approx:.1f}] s")

    # Phase 2: 精细二分搜索
    print("Phase 2: 精细二分搜索...")
    lo, hi = t_last_ok, t_coll_approx
    collision_info = None
    min_dist_at_collision = None

    for iteration in range(50):
        mid = (lo + hi) / 2

        # 首先检查几何可行性
        result, fail_handle = common.compute_dragon_at_time(mid)
        if result is None:
            hi = mid
            continue

        has_coll, min_dist, info = common.check_collision_at_time(mid)
        if has_coll:
            hi = mid
            collision_info = info
            min_dist_at_collision = min_dist
        else:
            lo = mid

        if hi - lo < tol:
            break

    t_collision = hi  # 第一个发生碰撞的时刻

    # 在碰撞时刻重新计算以获得完整数据
    theta_collision, _ = common.compute_dragon_at_time(t_collision)

    print(f"  碰撞时刻: t = {t_collision:.6f} s")
    if collision_info:
        print(f"  碰撞信息: {collision_info}, min_dist = {min_dist_at_collision:.6f} m")

    return t_collision, theta_collision, collision_info


def main():
    print("=" * 70)
    print("问题2: 确定盘入终止时刻 (碰撞检测)")
    print("=" * 70)

    # 搜索碰撞时刻 (从翻转后开始)
    t_term, theta_term, collision_info = find_first_collision(t_lo=300.0, t_hi=450.0)

    if t_term is None:
        print("\n未找到碰撞时刻！")
        return

    a = common.A_COEF

    # 终止时刻状态
    print(f"\n{'='*50}")
    print(f"终止时刻: t = {t_term:.6f} s")
    print(f"龙头前把手: θ = {theta_term[0]:.6f} ({theta_term[0]/np.pi:.4f}π), "
          f"r = {a * theta_term[0]:.4f} m")
    print(f"龙尾后把手: θ = {theta_term[-1]:.6f} ({theta_term[-1]/np.pi:.4f}π), "
          f"r = {a * theta_term[-1]:.4f} m")

    # 计算所有把手位置和速度
    positions = np.zeros((common.N_HANDLES, 2))
    for i in range(common.N_HANDLES):
        x, y = common.spiral_to_xy(theta_term[i], a)
        positions[i, 0] = x
        positions[i, 1] = y

    # 速度 (使用 t_term-dt 差分)
    dt = 0.001
    velocities = np.full(common.N_HANDLES, np.nan)
    if t_term >= dt:
        theta_before, _ = common.compute_dragon_at_time(t_term - dt)
        if theta_before is not None:
            for i in range(common.N_HANDLES):
                x_now, y_now = common.spiral_to_xy(theta_term[i], a)
                x_before, y_before = common.spiral_to_xy(theta_before[i], a)
                velocities[i] = np.sqrt((x_now - x_before)**2 + (y_now - y_before)**2) / dt

    # ---- 打印关键把手数据 ----
    key_labels = list(common.KEY_HANDLES.keys())
    key_indices = list(common.KEY_HANDLES.values())

    print(f"\n终止时刻关键把手位置和速度:")
    print(f"{'把手':>14s}  {'x (m)':>10s}  {'y (m)':>10s}  {'r (m)':>10s}  {'速度 (m/s)':>12s}")
    print("-" * 60)
    for label, ki in zip(key_labels, key_indices):
        r = a * theta_term[ki]
        print(f"{label:>14s}  {positions[ki,0]:10.6f}  {positions[ki,1]:10.6f}  "
              f"{r:10.4f}  {velocities[ki]:12.6f}")

    # 速度统计
    valid_vel = velocities[~np.isnan(velocities)]
    print(f"\n速度统计 (终止时刻):")
    print(f"  龙头速度: {velocities[0]:.4f} m/s")
    print(f"  尾部最大速度: {np.max(velocities[-20:]):.4f} m/s")
    print(f"  所有把手最大速度: {np.max(valid_vel):.4f} m/s")

    # ---- 碰撞详情 ----
    if collision_info:
        i, j, vname = collision_info
        print(f"\n碰撞详情:")
        print(f"  内侧板 Board[{i}]: handles {i}→{i+1}, "
              f"θ={theta_term[i]:.4f}→{theta_term[i+1]:.4f}")
        print(f"  外侧板 Board[{j}]: handles {j}→{j+1}, "
              f"θ={theta_term[j]:.4f}→{theta_term[j+1]:.4f}")
        print(f"  Δθ = {(theta_term[j] - theta_term[i])/np.pi:.4f}π")
        print(f"  r_i = {a*theta_term[i]:.4f}m, r_j = {a*theta_term[j]:.4f}m")
        print(f"  碰撞顶点: {vname}")

    # ---- 写入 result2.xlsx ----
    write_result2(t_term, theta_term, positions, velocities, collision_info)


def write_result2(t_term, theta_term, positions, velocities, collision_info):
    """Write result2.xlsx with termination state."""
    try:
        import openpyxl
    except ImportError:
        print("\n[WARNING] openpyxl not available.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '终止时刻'

    ws.cell(row=1, column=1, value=f'终止时刻: {t_term:.6f} s')
    if collision_info:
        ws.cell(row=2, column=1, value=f'碰撞对: Board[{collision_info[0]}] vs Board[{collision_info[1]}]')
    ws.cell(row=3, column=1, value='')

    headers = ['把手', 'theta (rad)', 'r (m)', 'x (m)', 'y (m)', '速度 (m/s)']
    for j, h in enumerate(headers):
        ws.cell(row=4, column=1 + j, value=h)

    a = common.A_COEF
    for i in range(common.N_HANDLES):
        row = 5 + i
        if i == 0:
            name = '龙头'
        elif i == common.N_HANDLES - 1:
            name = '龙尾（后）'
        elif i == common.N_HANDLES - 2:
            name = '龙尾'
        else:
            name = f'第{i}节龙身'

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(float(theta_term[i]), 6))
        ws.cell(row=row, column=3, value=round(float(a * theta_term[i]), 6))
        ws.cell(row=row, column=4, value=round(float(positions[i, 0]), 6))
        ws.cell(row=row, column=5, value=round(float(positions[i, 1]), 6))
        ws.cell(row=row, column=6, value=round(float(velocities[i]), 6))

    filepath = OUTPUT_DIR / 'result2.xlsx'
    wb.save(str(filepath))
    print(f"\n结果已保存到: {filepath}")


if __name__ == '__main__':
    main()
