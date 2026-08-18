"""
问题1: 舞龙队沿等距螺线盘入，300秒运动学模拟。
输出: result1.xlsx (位置 + 速度两个sheet)
"""
import numpy as np
from pathlib import Path
import common

OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)


def format_table_row(label, values, fmt='.6f'):
    """Format a table row with aligned columns."""
    return f"{label:>14s}" + "".join(f"{v:{fmt}}" if not np.isnan(v) else f"{'--':>12s}"
                                      for v in values)


def main():
    print("=" * 70)
    print("问题1: 板凳龙盘入运动学模拟 (0-300 s)")
    print("=" * 70)

    # Run simulation
    t_array, positions, velocities, theta_all, failed_at = common.simulate_time_range(0, 300)

    n_valid = np.sum(~np.isnan(positions[:, 0, 0]))
    print(f"有效时间步: {n_valid}/{len(t_array)}")

    if failed_at:
        t_idx, t_fail, fail_handle = failed_at
        print(f"\n警告: 在 t = {t_fail:.1f} s 时把手 #{fail_handle} (龙尾后把手) 无法定位")
        print("原因: 螺线内侧半径过小，无法容纳 1.65m 板凳间距。")
        print("这恰好对应问题2的终止条件。\n")

    # ---- Print key-time tables ----
    key_times = [0, 60, 120, 180, 240, 300]
    key_labels = list(common.KEY_HANDLES.keys())
    key_indices = list(common.KEY_HANDLES.values())

    print("\n" + "=" * 70)
    print("表1: 位置 (m)")
    print("=" * 70)
    header = f"{'':>14s}" + "".join(f"{t:>12d}s" for t in key_times)
    print(header)
    print("-" * (14 + 12 * len(key_times)))

    for label, ki in zip(key_labels, key_indices):
        for coord in ['x', 'y']:
            vals = []
            for t in key_times:
                t_idx = int(t)  # dt=1
                if t_idx < n_valid:
                    v = positions[t_idx, ki, 0 if coord == 'x' else 1]
                else:
                    v = np.nan
                vals.append(v)
            print(format_table_row(f"{label} {coord}", vals))

    print("\n" + "=" * 70)
    print("表2: 速度 (m/s)")
    print("=" * 70)
    print(header)
    print("-" * (14 + 12 * len(key_times)))

    for label, ki in zip(key_labels, key_indices):
        vals = []
        for t in key_times:
            t_idx = int(t)
            if t_idx < n_valid:
                v = velocities[t_idx, ki]
            else:
                v = np.nan
            vals.append(v)
        print(format_table_row(label, vals))

    # ---- Speed statistics ----
    valid_vel = velocities[:n_valid, :]
    print(f"\n速度统计:")
    print(f"  所有把手最大速度: {np.nanmax(valid_vel):.4f} m/s")
    print(f"  所有把手最小速度: {np.nanmin(valid_vel):.4f} m/s")
    print(f"  龙头平均速度:     {np.nanmean(valid_vel[:, 0]):.4f} m/s")
    print(f"  龙尾最大速度:     {np.nanmax(valid_vel[:, -1]):.4f} m/s")

    # ---- Write result1.xlsx ----
    write_result1(t_array, positions, velocities, theta_all, n_valid)


def write_result1(t_array, positions, velocities, theta_all, n_valid):
    """Write result1.xlsx with position and velocity sheets."""
    try:
        import openpyxl
    except ImportError:
        print("\n[WARNING] openpyxl not available, skipping Excel output.")
        return

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 位置 ----
    ws_pos = wb.active
    ws_pos.title = '位置'

    # Header row: time labels
    ws_pos.cell(row=1, column=1, value='')
    for j in range(n_valid):
        ws_pos.cell(row=1, column=2 + j, value=f'{int(t_array[j])} s')

    # Data rows: for each handle, one row for x, one for y
    handle_names = ['龙头'] + [f'第{i}节龙身' for i in range(1, 222)] + ['龙尾', '龙尾（后）']
    # Actually we have 224 handles: 0=head front, 1=head rear/body1 front, ..., 222=tail front, 223=tail rear

    row = 2
    for i in range(common.N_HANDLES):
        # Determine name
        if i == 0:
            name = '龙头'
        elif i == common.N_HANDLES - 1:
            name = '龙尾（后）'
        elif i == common.N_HANDLES - 2:
            name = '龙尾'
        else:
            name = f'第{i}节龙身'

        # x row
        ws_pos.cell(row=row, column=1, value=f'{name} x (m)')
        for j in range(n_valid):
            ws_pos.cell(row=row, column=2 + j, value=round(float(positions[j, i, 0]), 6))
        row += 1

        # y row
        ws_pos.cell(row=row, column=1, value=f'{name} y (m)')
        for j in range(n_valid):
            ws_pos.cell(row=row, column=2 + j, value=round(float(positions[j, i, 1]), 6))
        row += 1

    # ---- Sheet 2: 速度 ----
    ws_vel = wb.create_sheet('速度')

    ws_vel.cell(row=1, column=1, value='')
    for j in range(n_valid):
        ws_vel.cell(row=1, column=2 + j, value=f'{int(t_array[j])} s')

    row = 2
    for i in range(common.N_HANDLES):
        if i == 0:
            name = '龙头'
        elif i == common.N_HANDLES - 1:
            name = '龙尾（后）'
        elif i == common.N_HANDLES - 2:
            name = '龙尾'
        else:
            name = f'第{i}节龙身'

        ws_vel.cell(row=row, column=1, value=f'{name} (m/s)')
        for j in range(n_valid):
            ws_vel.cell(row=row, column=2 + j, value=round(float(velocities[j, i]), 6))
        row += 1

    filepath = OUTPUT_DIR / 'result1.xlsx'
    wb.save(str(filepath))
    print(f"\n结果已保存到: {filepath}")


if __name__ == '__main__':
    main()
